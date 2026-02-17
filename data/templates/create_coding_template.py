#!/usr/bin/env python3
"""
Generate Excel Coding Template for Agentic AI Learning Outcomes Meta-Analysis
Created: 2026-02-16
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def auto_fit_columns(sheet):
    """Auto-fit column widths based on content"""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        sheet.column_dimensions[column_letter].width = adjusted_width

def create_header_row(sheet, headers):
    """Create formatted header row"""
    header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Freeze top row
    sheet.freeze_panes = "A2"

def add_dropdown_validation(sheet, col_letter, start_row, options, allow_blank=True):
    """Add data validation dropdown to a column"""
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=allow_blank)
    dv.error = 'Please select from the dropdown list'
    dv.errorTitle = 'Invalid Entry'
    sheet.add_data_validation(dv)
    dv.add(f'{col_letter}{start_row}:{col_letter}1000')

def create_study_metadata_sheet(wb):
    """Sheet 1: Study_Metadata"""
    ws = wb.create_sheet("Study_Metadata", 0)
    ws.sheet_properties.tabColor = "4472C4"

    headers = [
        "study_id", "authors", "year", "title", "journal", "doi", "country",
        "n_treatment", "n_control", "n_total", "pct_female", "mean_age",
        "pub_type", "study_design", "random_assignment", "duration_weeks", "notes"
    ]
    create_header_row(ws, headers)

    # Add dropdowns
    add_dropdown_validation(ws, 'M', 2, ["journal", "conference", "dissertation", "preprint"])
    add_dropdown_validation(ws, 'N', 2, ["RCT", "quasi_experimental", "pre_post", "single_subject"])
    add_dropdown_validation(ws, 'O', 2, ["yes", "no", "NA"])

    auto_fit_columns(ws)

def create_effect_sizes_sheet(wb):
    """Sheet 2: Effect_Sizes"""
    ws = wb.create_sheet("Effect_Sizes", 1)
    ws.sheet_properties.tabColor = "70AD47"

    headers = [
        "study_id", "es_id", "outcome_measure", "outcome_type", "outcome_level_blooms",
        "measurement_type", "timing", "treatment_m", "treatment_sd", "treatment_n",
        "control_m", "control_sd", "control_n", "pre_m", "pre_sd", "post_m", "post_sd",
        "hedges_g", "se_g", "ci_lower", "ci_upper", "es_source", "original_statistic",
        "original_value", "notes"
    ]
    create_header_row(ws, headers)

    # Add dropdowns
    add_dropdown_validation(ws, 'D', 2, ["cognitive", "skill_based", "affective", "performance"])
    add_dropdown_validation(ws, 'E', 2, ["remember_understand", "apply_analyze", "evaluate_create"])
    add_dropdown_validation(ws, 'F', 2, ["standardized_test", "researcher_developed", "performance_assessment", "self_report"])
    add_dropdown_validation(ws, 'G', 2, ["immediate_post", "delayed_post", "transfer"])
    add_dropdown_validation(ws, 'V', 2, ["computed_from_means", "reported_d", "reported_g", "computed_from_F", "computed_from_t", "computed_from_r"])

    auto_fit_columns(ws)

def create_ai_agent_characteristics_sheet(wb):
    """Sheet 3: AI_Agent_Characteristics"""
    ws = wb.create_sheet("AI_Agent_Characteristics", 2)
    ws.sheet_properties.tabColor = "FFC000"

    headers = [
        "study_id", "agent_name_description", "human_oversight_level", "agent_architecture",
        "num_agents", "agency_level_apcp", "agent_role", "agent_modality", "ai_technology",
        "adaptivity", "feedback_type", "personalization", "notes"
    ]
    create_header_row(ws, headers)

    # Add dropdowns
    add_dropdown_validation(ws, 'C', 2, ["fully_autonomous", "ai_led_checkpoints", "human_led_ai_support"])
    add_dropdown_validation(ws, 'D', 2, ["single_agent", "multi_agent"])
    add_dropdown_validation(ws, 'F', 2, ["adaptive", "proactive", "co_learner", "peer"])
    add_dropdown_validation(ws, 'G', 2, ["tutor", "coach", "assessor", "collaborator", "facilitator"])
    add_dropdown_validation(ws, 'H', 2, ["text_only", "voice", "embodied_avatar", "mixed"])
    add_dropdown_validation(ws, 'I', 2, ["rule_based", "ML", "NLP", "LLM", "reinforcement_learning"])
    add_dropdown_validation(ws, 'J', 2, ["static", "adaptive_performance", "adaptive_behavior_affect"])
    add_dropdown_validation(ws, 'K', 2, ["immediate", "delayed", "on_demand", "none"])
    add_dropdown_validation(ws, 'L', 2, ["none", "content", "pace", "both"])

    auto_fit_columns(ws)

def create_learning_context_sheet(wb):
    """Sheet 4: Learning_Context"""
    ws = wb.create_sheet("Learning_Context", 3)
    ws.sheet_properties.tabColor = "9966FF"

    headers = [
        "study_id", "learning_context", "education_level", "domain", "specific_subject",
        "learning_mode", "delivery", "comparison_condition", "treatment_duration_sessions", "notes"
    ]
    create_header_row(ws, headers)

    # Add dropdowns
    add_dropdown_validation(ws, 'B', 2, ["k12", "higher_education", "workplace_training", "professional_education", "continuing_education"])
    add_dropdown_validation(ws, 'C', 2, ["elementary", "middle", "high_school", "undergraduate", "graduate", "adult_professional"])
    add_dropdown_validation(ws, 'D', 2, ["STEM", "language", "medical", "business", "ICT", "arts", "other"])
    add_dropdown_validation(ws, 'F', 2, ["formal", "informal", "blended"])
    add_dropdown_validation(ws, 'G', 2, ["face_to_face_with_AI", "fully_online", "hybrid"])
    add_dropdown_validation(ws, 'H', 2, ["no_AI", "non_agentic_AI", "human_only", "traditional_instruction"])

    auto_fit_columns(ws)

def create_quality_assessment_sheet(wb):
    """Sheet 5: Quality_Assessment"""
    ws = wb.create_sheet("Quality_Assessment", 4)
    ws.sheet_properties.tabColor = "FF6666"

    headers = [
        "study_id", "random_assignment_quality", "sample_representativeness",
        "measurement_validity", "attrition_bias", "reporting_completeness",
        "treatment_fidelity", "overall_risk_of_bias", "exclude_from_sensitivity", "notes"
    ]
    create_header_row(ws, headers)

    # Add dropdowns
    risk_levels = ["low_risk", "some_concerns", "high_risk"]
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        add_dropdown_validation(ws, col_letter, 2, risk_levels)

    add_dropdown_validation(ws, 'I', 2, ["TRUE", "FALSE"])

    auto_fit_columns(ws)

def create_ai_extraction_provenance_sheet(wb):
    """Sheet 6: AI_Extraction_Provenance"""
    ws = wb.create_sheet("AI_Extraction_Provenance", 5)
    ws.sheet_properties.tabColor = "00B0F0"

    headers = [
        "study_id", "extraction_phase", "ai_model_primary", "extraction_timestamp",
        "confidence_score", "consensus_agreement", "human_verified", "discrepancy_notes", "cost_usd"
    ]
    create_header_row(ws, headers)

    # Add dropdowns
    add_dropdown_validation(ws, 'B', 2, ["initial_screening", "full_text_review", "data_extraction", "quality_assessment"])
    add_dropdown_validation(ws, 'C', 2, ["claude-opus-4", "gpt-4", "gemini-pro", "human_only"])
    add_dropdown_validation(ws, 'G', 2, ["yes", "no", "pending"])

    auto_fit_columns(ws)

def main():
    """Generate the Excel coding template"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Create all sheets
    create_study_metadata_sheet(wb)
    create_effect_sizes_sheet(wb)
    create_ai_agent_characteristics_sheet(wb)
    create_learning_context_sheet(wb)
    create_quality_assessment_sheet(wb)
    create_ai_extraction_provenance_sheet(wb)

    # Save workbook
    output_path = "Agentic_AI_Learning_MA_Coding_v1.xlsx"
    wb.save(output_path)
    print(f"✓ Created: {output_path}")
    print(f"  - {len(wb.sheetnames)} sheets")
    print(f"  - Sheets: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
