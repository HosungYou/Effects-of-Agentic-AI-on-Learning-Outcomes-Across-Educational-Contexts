#!/usr/bin/env python3
"""
Generate Excel coding template for manual effect size extraction.
Creates a structured spreadsheet with validation, instructions, and drop-down lists.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl required: pip install openpyxl")


# Colour palette
COLOURS = {
    'header_dark':   "2C3E50",
    'header_blue':   "2980B9",
    'header_green':  "27AE60",
    'header_orange': "E67E22",
    'header_purple': "8E44AD",
    'row_alt':       "EBF5FB",
    'row_white':     "FFFFFF",
    'required':      "FDFEFE",
    'optional':      "F8F9FA",
    'instruction':   "FEF9E7"
}

# Drop-down options
DROPDOWNS = {
    'design_type': [
        'randomized_controlled_trial', 'quasi_experimental',
        'pre_post', 'other'
    ],
    'education_level': [
        'k12', 'higher_education', 'adult', 'professional', 'mixed'
    ],
    'outcome_type': [
        'learning_outcome', 'achievement', 'test_score', 'knowledge',
        'skill', 'competency', 'grade', 'satisfaction', 'engagement', 'other'
    ],
    'data_format': [
        'between_subjects_MSD', 'within_subjects_prepost',
        'cohens_d', 'hedges_g', 't_statistic', 'F_statistic', 'other'
    ],
    'oversight_level': [
        'fully_autonomous', 'ai_led_with_checkpoints',
        'human_led_with_ai_support'
    ],
    'architecture': ['single_agent', 'multi_agent'],
    'agency_level': ['adaptive', 'proactive', 'co_learner', 'peer'],
    'role': ['tutor', 'coach', 'assessor', 'collaborator', 'facilitator'],
    'modality': ['text_only', 'voice', 'embodied', 'mixed'],
    'technology': ['rule_based', 'ml', 'nlp', 'llm', 'rl'],
    'adaptivity': [
        'static', 'adaptive_performance', 'adaptive_behavior_affect'
    ],
    'confidence': ['high', 'moderate', 'low']
}


def _header_style(colour: str, bold: bool = True, font_colour: str = "FFFFFF",
                  size: int = 11) -> dict:
    return {
        'fill': PatternFill("solid", fgColor=colour),
        'font': Font(bold=bold, color=font_colour, size=size),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def _apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int,
                  options: List[str]):
    formula = '"' + ','.join(options) + '"'
    dv = DataValidation(
        type='list',
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='Invalid value',
        error='Please select from the drop-down list.'
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def create_coding_template(output_path: str = "data/coding_template.xlsx",
                           n_rows: int = 200):
    """
    Create Excel coding template with structured sheets.

    Args:
        output_path: Path to save the template
        n_rows: Number of data rows to pre-format
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _create_instructions_sheet(wb)
    _create_study_info_sheet(wb, n_rows)
    _create_effect_sizes_sheet(wb, n_rows)
    _create_agent_characteristics_sheet(wb, n_rows)
    _create_codebook_sheet(wb)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    logger.info(f"Coding template saved: {output}")
    print(f"Coding template created: {output}")


def _create_instructions_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("00_INSTRUCTIONS")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80

    title_style = _header_style(COLOURS['header_dark'], size=14)
    _apply_style(ws['A1'], title_style)
    _apply_style(ws['B1'], title_style)
    ws['A1'] = "FIELD"
    ws['B1'] = "AGENTIC AI LEARNING OUTCOMES META-ANALYSIS — CODING TEMPLATE INSTRUCTIONS"
    ws.row_dimensions[1].height = 30

    instructions = [
        ("Purpose", "This template is for extracting effect size and study characteristic data "
         "from included studies. Each study gets ONE row in sheets 01 and 03. "
         "Multiple effect sizes from the same study each get their own row in sheet 02."),
        ("Sheet 01", "Study-level information: design, sample, context. One row per study."),
        ("Sheet 02", "Effect size data: one row per outcome measure per study. "
         "Link to Sheet 01 via study_id."),
        ("Sheet 03", "AI agent characteristics: one row per study."),
        ("Required fields", "Columns with * in header are REQUIRED. Leave blank only if "
         "truly not reported in the paper."),
        ("Drop-downs", "Columns with coloured headers have drop-down validation. "
         "Select from the list; do not type free text."),
        ("study_id", "Assign a unique ID to each study: e.g., Smith2023, Jones2024b. "
         "Must be consistent across all three sheets."),
        ("Confidence", "Your subjective confidence in the extracted value: "
         "high = explicitly stated, moderate = inferable, low = uncertain."),
        ("Notes", "Use the notes column for any extraction caveats, page references, "
         "or ambiguities."),
        ("Version", f"Template created: {datetime.now().strftime('%Y-%m-%d')}")
    ]

    inst_style = {
        'font': Font(size=10),
        'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
        'border': Border(bottom=Side(style='thin'))
    }

    for i, (field, text) in enumerate(instructions, start=2):
        ws[f'A{i}'] = field
        ws[f'B{i}'] = text
        _apply_style(ws[f'A{i}'], {**inst_style, 'font': Font(bold=True, size=10)})
        _apply_style(ws[f'B{i}'], inst_style)
        ws.row_dimensions[i].height = 45

    ws.freeze_panes = 'A2'


def _create_study_info_sheet(wb: openpyxl.Workbook, n_rows: int):
    ws = wb.create_sheet("01_Study_Info")
    ws.freeze_panes = 'C2'

    columns = [
        # (header, width, colour, dropdown_key, required)
        ("study_id *",              16, COLOURS['header_dark'],   None,             True),
        ("first_author *",          20, COLOURS['header_dark'],   None,             True),
        ("year *",                  8,  COLOURS['header_dark'],   None,             True),
        ("title",                   40, COLOURS['header_dark'],   None,             False),
        ("journal",                 30, COLOURS['header_dark'],   None,             False),
        ("doi",                     30, COLOURS['header_dark'],   None,             False),
        ("design_type *",           22, COLOURS['header_blue'],   'design_type',    True),
        ("has_control_group",       18, COLOURS['header_blue'],   None,             False),
        ("n_treatment *",           14, COLOURS['header_blue'],   None,             True),
        ("n_control",               12, COLOURS['header_blue'],   None,             False),
        ("n_total *",               12, COLOURS['header_blue'],   None,             True),
        ("education_level *",       18, COLOURS['header_green'],  'education_level',True),
        ("learning_domain *",       22, COLOURS['header_green'],  None,             True),
        ("country *",               16, COLOURS['header_green'],  None,             True),
        ("duration_weeks",          14, COLOURS['header_green'],  None,             False),
        ("outcome_type *",          22, COLOURS['header_orange'], 'outcome_type',   True),
        ("outcome_measure_desc",    35, COLOURS['header_orange'], None,             False),
        ("pretest_used",            14, COLOURS['header_orange'], None,             False),
        ("confidence *",            14, COLOURS['header_purple'], 'confidence',     True),
        ("notes",                   40, COLOURS['header_dark'],   None,             False),
    ]

    for col_idx, (header, width, colour, dropdown, _) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_style(cell, _header_style(colour))
        if dropdown and n_rows > 0:
            _add_dropdown(ws, letter, 2, n_rows + 1, DROPDOWNS[dropdown])

    ws.row_dimensions[1].height = 35

    for row in range(2, n_rows + 2):
        fill_colour = COLOURS['row_alt'] if row % 2 == 0 else COLOURS['row_white']
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_colour)


def _create_effect_sizes_sheet(wb: openpyxl.Workbook, n_rows: int):
    ws = wb.create_sheet("02_Effect_Sizes")
    ws.freeze_panes = 'C2'

    columns = [
        ("study_id *",             16, COLOURS['header_dark'],   None,          True),
        ("outcome_label *",        30, COLOURS['header_dark'],   None,          True),
        ("outcome_type *",         22, COLOURS['header_orange'], 'outcome_type',True),
        ("data_format *",          25, COLOURS['header_blue'],   'data_format', True),
        # Between-subjects
        ("m_treatment",            14, COLOURS['header_blue'],   None,          False),
        ("sd_treatment",           14, COLOURS['header_blue'],   None,          False),
        ("n_treatment",            14, COLOURS['header_blue'],   None,          False),
        ("m_control",              14, COLOURS['header_blue'],   None,          False),
        ("sd_control",             14, COLOURS['header_blue'],   None,          False),
        ("n_control",              14, COLOURS['header_blue'],   None,          False),
        # Pre-post
        ("m_pre",                  12, COLOURS['header_green'],  None,          False),
        ("sd_pre",                 12, COLOURS['header_green'],  None,          False),
        ("m_post",                 12, COLOURS['header_green'],  None,          False),
        ("sd_post",                12, COLOURS['header_green'],  None,          False),
        ("n_prepost",              12, COLOURS['header_green'],  None,          False),
        # Precomputed
        ("cohens_d",               12, COLOURS['header_orange'], None,          False),
        ("hedges_g",               12, COLOURS['header_orange'], None,          False),
        ("se_g",                   12, COLOURS['header_orange'], None,          False),
        ("t_statistic",            14, COLOURS['header_orange'], None,          False),
        ("F_statistic",            14, COLOURS['header_orange'], None,          False),
        ("p_value",                12, COLOURS['header_orange'], None,          False),
        ("source_table_page",      20, COLOURS['header_dark'],   None,          False),
        ("confidence *",           14, COLOURS['header_purple'], 'confidence',  True),
        ("notes",                  40, COLOURS['header_dark'],   None,          False),
    ]

    for col_idx, (header, width, colour, dropdown, _) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_style(cell, _header_style(colour))
        if dropdown and n_rows > 0:
            _add_dropdown(ws, letter, 2, n_rows + 1, DROPDOWNS[dropdown])

    ws.row_dimensions[1].height = 35

    for row in range(2, n_rows + 2):
        fill_colour = COLOURS['row_alt'] if row % 2 == 0 else COLOURS['row_white']
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_colour)


def _create_agent_characteristics_sheet(wb: openpyxl.Workbook, n_rows: int):
    ws = wb.create_sheet("03_Agent_Characteristics")
    ws.freeze_panes = 'B2'

    columns = [
        ("study_id *",        16, COLOURS['header_dark'],   None,             True),
        ("ai_system_name",    25, COLOURS['header_dark'],   None,             False),
        ("oversight_level *", 25, COLOURS['header_blue'],   'oversight_level',True),
        ("architecture *",    18, COLOURS['header_blue'],   'architecture',   True),
        ("agency_level *",    18, COLOURS['header_blue'],   'agency_level',   True),
        ("role *",            18, COLOURS['header_green'],  'role',           True),
        ("modality *",        16, COLOURS['header_green'],  'modality',       True),
        ("technology *",      16, COLOURS['header_orange'], 'technology',     True),
        ("adaptivity *",      25, COLOURS['header_orange'], 'adaptivity',     True),
        ("confidence *",      14, COLOURS['header_purple'], 'confidence',     True),
        ("notes",             40, COLOURS['header_dark'],   None,             False),
    ]

    for col_idx, (header, width, colour, dropdown, _) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_style(cell, _header_style(colour))
        if dropdown and n_rows > 0:
            _add_dropdown(ws, letter, 2, n_rows + 1, DROPDOWNS[dropdown])

    ws.row_dimensions[1].height = 35

    for row in range(2, n_rows + 2):
        fill_colour = COLOURS['row_alt'] if row % 2 == 0 else COLOURS['row_white']
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_colour)


def _create_codebook_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("04_Codebook")
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60

    headers = ["Variable", "Code", "Definition"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, _header_style(COLOURS['header_dark']))

    entries = []
    for field, options in DROPDOWNS.items():
        for opt in options:
            entries.append((field, opt, ""))

    for row_idx, (field, code, defn) in enumerate(entries, start=2):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=code)
        ws.cell(row=row_idx, column=3, value=defn)
        fill = COLOURS['row_alt'] if row_idx % 2 == 0 else COLOURS['row_white']
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=fill)


if __name__ == "__main__":
    import argparse

    logging.basicConfig(level=logging.INFO, format='%(levelname)s - %(message)s')

    parser = argparse.ArgumentParser(
        description="Generate Excel coding template for meta-analysis"
    )
    parser.add_argument('--output', default='data/coding_template.xlsx',
                        help='Output path for the Excel template')
    parser.add_argument('--rows', type=int, default=200,
                        help='Number of pre-formatted data rows (default: 200)')

    args = parser.parse_args()
    create_coding_template(args.output, args.rows)
